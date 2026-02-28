import os
from datetime import date
from openpyxl import load_workbook
from copy import copy

def _s(v):
    return "" if v is None else str(v).strip()

def _upper(v):
    return _s(v).upper()

def _pick_item(items):
    # Keep only dict items
    items = [x for x in (items or []) if isinstance(x, dict)]

    # Prefer DG items, else fallback to first item
    for it in items:
        if _upper(it.get("dgStatus")) in ("DG", "YES"):
            return it
    return items[0] if items else {}

def _wrap(ws, cell_addr):
    c = ws[cell_addr]
    c.alignment = copy(c.alignment)
    c.alignment = c.alignment.copy(wrap_text=True)

def generate_dg_excel(template_path: str, output_path: str, payload: dict):
    wb = load_workbook(template_path)
    ws = wb["DG Form"] if "DG Form" in wb.sheetnames else wb.active

    items = payload.get("items") or []
    it = _pick_item(items)

    # --- date (you said YES). We'll write into the template cell that already exists for Date.
    # Your earlier screenshot shows the "Date" value box at top-right.
    # If your template cell for the date is fixed, put it here. If not, we keep it simple:
    # We'll set the cell that currently holds the date value if it exists, else we set F? only if you tell me.
    #
    # FAST approach: set today's date into the common box used in your sample.
    # If it doesn't land correctly, tell me the exact cell for the Date field and I'll lock it.
    #
    # For now: we set it to the same place you want other fixed fields: use label search is more complex.
    # We'll use a safe default: try G8 (often used), fallback do nothing.
    from datetime import date
    today = date.today().strftime("%B %d, %Y").upper()
    ws["D25"].value = _upper(it.get("description"))                 # UNNO

    ws["G9"].value = today
    # --- F26..F39 fields (as your list)
    ws["F26"].value = _upper(it.get("unNumber"))                 # UNNO
    ws["F27"].value = _upper(it.get("classNumber"))              # IMCO CLASS
    ws["F28"].value = _upper(it.get("packingGroup"))             # PACKING GROUP
    ws["F29"].value = _upper(it.get("flashPoint"))               # FLASH POINT
    ws["F30"].value = _upper(it.get("outerType"))                # OUTER PACKAGE TYPE
    ws["F31"].value = "4G"                                       # PACKING CODE (always)
    ws["F33"].value = _upper(it.get("innerType"))                # INNER PACKAGE TYPE
    # F34 INNER PACKAGE QUANTITY = NO (leave blank)
    ws["F35"].value = _upper(it.get("technicalName"))            # TECHNICAL NAME
    ws["F36"].value = _upper(it.get("properShippingName"))       # PROPER SHIPPING NAME
    ws["F37"].value = "N/A"                                      # SEGREGATION GROUP (always)
    ws["F38"].value = _upper(it.get("marinePollutant"))          # MARINE POLLUTANT
    ws["F39"].value = _upper(it.get("ems"))                      # EMS F-S
    ws["G47"].value = "4G"                                       # PACKING CODE (always)

    # --- weights
    ws["H27"].value = _upper(it.get("grossWeight"))              # gross weight
    ws["H28"].value = _upper(it.get("netWeight"))                # net weight

    # --- bottom summary row
    ws["B47"].value = _upper(it.get("classNumber"))              # IMO CLASS
    ws["D47"].value = _upper(it.get("unNumber"))                 # UN NUM
    ws["E47"].value = _upper(it.get("packingGroup"))             # packing group
    ws["F47"].value = _upper(it.get("flashPoint"))               # flashpoint

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path